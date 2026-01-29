import os
import boto3
from werkzeug.utils import secure_filename
from openpyxl import Workbook, load_workbook
from datetime import datetime
import re
import sys
from aws_config import AWS_ACCESS_KEY, AWS_SECRET_KEY, AWS_REGION

ALLOWED_EXTENSIONS = {'.jpg', '.jpeg', '.png'}
MAX_FILE_SIZE_MB = 5
EXCEL_FILE = 'students.xlsx'
BUCKET_NAME = 'ict-attendances'

s3 = boto3.client(
    's3',
    region_name=AWS_REGION,
    aws_access_key_id=AWS_ACCESS_KEY,
    aws_secret_access_key=AWS_SECRET_KEY
)

def allowed_file(filename):
    _, ext = os.path.splitext(filename)
    return ext.lower() in ALLOWED_EXTENSIONS

def file_size_okay(file_obj):
    file_obj.seek(0, os.SEEK_END)
    size_mb = file_obj.tell() / (1024 * 1024)
    file_obj.seek(0)
    return size_mb <= MAX_FILE_SIZE_MB

def upload_file_to_s3(bucket_name, file_path, s3_key):
    s3.upload_file(file_path, bucket_name, s3_key)

def sanitize_for_s3_key(text: str) -> str:
    text = text.strip().replace(" ", "_")
    return re.sub(r'[^a-zA-Z0-9_\-]', '', text)

def update_student_excel(batch_name, er_number, name, parent_phone):
    if os.path.exists(EXCEL_FILE):
        wb = load_workbook(EXCEL_FILE)
    else:
        wb = Workbook()

    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    if batch_name not in wb.sheetnames:
        batch_sheet = wb.create_sheet(batch_name)
        batch_sheet.append(
            ["ER Number", "Student Name", "Parent Phone", "Batch Name", "Upload Date & Time"]
        )
    else:
        batch_sheet = wb[batch_name]

    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    batch_sheet.append([er_number, name, parent_phone, batch_name, now])

    summary_name = "Batch Info"
    if summary_name not in wb.sheetnames:
        summary = wb.create_sheet(summary_name)
        summary.append(
            ["Batch Name", "ER Number", "Student Name", "Parent Phone", "Last Updated"]
        )
    else:
        summary = wb[summary_name]

    existing = [(r[0].value, r[1].value) for r in summary.iter_rows(min_row=2)]
    if (batch_name, er_number) not in existing:
        summary.append([batch_name, er_number, name, parent_phone, now])

    wb.save(EXCEL_FILE)

def upload_multiple_images(batch_name, er_number, name, parent_phone, image_files):
    er_number = er_number.strip()
    sanitized_batch = sanitize_for_s3_key(batch_name)
    sanitized_name = sanitize_for_s3_key(name)

    os.makedirs("uploads", exist_ok=True)
    results = []

    for i, image_file in enumerate(image_files):
        filename = secure_filename(image_file.filename)
        ext = os.path.splitext(filename)[1].lower()

        if not allowed_file(filename):
            results.append(f"Rejected {filename}")
            continue

        if not file_size_okay(image_file):
            results.append(f"Rejected {filename}: too large")
            continue

        new_filename = f"{er_number}_{sanitized_name}_{i+1}{ext}"
        s3_key = f"{sanitized_batch}/{new_filename}"
        local_path = os.path.join("uploads", new_filename)

        image_file.save(local_path)
        s3.upload_file(local_path, BUCKET_NAME, s3_key)
        results.append(f"✅ Uploaded: {s3_key}")

        index_face_to_rekognition(er_number, sanitized_name, s3_key)
        os.remove(local_path)

    update_student_excel(batch_name, er_number, name, parent_phone)
    upload_file_to_s3(BUCKET_NAME, EXCEL_FILE, EXCEL_FILE)
    results.append("✅ Excel updated & uploaded")

    return results

def index_face_to_rekognition(er_number, student_name, s3_key, collection_id="students"):
    rekognition = boto3.client("rekognition", region_name=AWS_REGION)
    external_id = f"{er_number}_{student_name}"

    try:
        rekognition.index_faces(
            CollectionId=collection_id,
            Image={"S3Object": {"Bucket": BUCKET_NAME, "Name": s3_key}},
            ExternalImageId=external_id
        )
    except rekognition.exceptions.ResourceNotFoundException:
        rekognition.create_collection(CollectionId=collection_id)
        index_face_to_rekognition(er_number, student_name, s3_key, collection_id)
